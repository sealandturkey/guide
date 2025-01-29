import csv
from django.http import HttpResponse
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from .models import CustomUser, PreRegisteredUser
from admin_extra_buttons.api import ExtraButtonsMixin, button  # Gerekli eklenti
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class CustomUserAdmin(UserAdmin, ExtraButtonsMixin):
    # Admin panelinde gösterilecek alanlar
    list_display = ('email', 'first_name', 'last_name', 'phone_number')
    search_fields = ('email', 'first_name', 'last_name', 'phone_number')
    ordering = ('email',)

    # Kullanıcı detay sayfasında gösterilecek alanlar
    fieldsets = (
        (None, {'fields': ('email', 'password')}),
        ('Personal Info', {'fields': ('first_name', 'last_name', 'phone_number')}),
        # ('Permissions', {'fields': ('is_active', 'is_staff', 'is_superuser')}),
    )

    # Kullanıcı oluşturma formu için gerekli alanlar
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('email', 'first_name', 'last_name', 'phone_number', 'password1', 'password2', 'is_active', 'is_staff', 'is_superuser'),
        }),
    )
    
    # Kullanıcı oluştururken ve giriş yaparken `username` alanını gizle
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'username' in form.base_fields:
            form.base_fields['username'].widget = admin.widgets.HiddenInput()
        return form

    # Extra button: Export all users to Excel
    @button(label="Export to CSV", html_attrs={'class': 'btn btn-primary'})  # Eklenti butonu
    def export_to_csv(self, request):
        # HTTP Response (Excel dosyası için)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=custom_users.csv'

        # CSV Writer
        writer = csv.writer(response)
        writer.writerow(['Email', 'First Name', 'Last Name', 'Phone Number'])  # Başlık satırı

        # Tüm kullanıcıları yaz
        for user in CustomUser.objects.all():
            writer.writerow([user.first_name, user.last_name, user.email, user.phone_number])

        return response  # CSV dosyasını döndür

    @button(label="Export to Excel", html_attrs={'class': 'btn btn-success'})  # Eklenti butonu
    def export_to_excel(self, request):
        # Yeni bir Excel Workbook oluştur
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Custom Users"

        # Başlık satırını ekle
        headers = ['Email', 'First Name', 'Last Name', 'Phone Number']
        worksheet.append(headers)

        # Tüm kullanıcıların verilerini yaz
        for user in CustomUser.objects.all():
            worksheet.append([user.email, user.first_name, user.last_name, user.phone_number])

        # HTTP Response ayarları (Excel dosyası için)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=custom_users.xlsx'

        # Sütun boyutlarını ayarla
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)  # Sütun harfini al (A, B, C, ...)
            worksheet.column_dimensions[col_letter].width = 30  # Sütun genişliğini 20 olarak ayarla
        
        # Excel dosyasını response'a kaydet
        workbook.save(response)
        return response

# CustomUser modelini admin paneline ekle
admin.site.register(CustomUser, CustomUserAdmin)
admin.site.register(PreRegisteredUser)